import re
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

current_dir = Path(__file__).parent
rtf_files = list(current_dir.glob("*.rtf"))

NUMBER_PATTERN = r"[-+]?\d*\.\d+|[-+]?\d+"


def rtf_to_text(rtf: str) -> str:
    text = rtf.replace("\r", "")
    text = re.sub(r"\\par[d]? ?|\\line\b ?", "\n", text)

    def decode_hex(match):
        try:
            return bytes.fromhex(match.group(1)).decode("cp1252")
        except Exception:
            return ""

    text = re.sub(r"\\'([0-9a-fA-F]{2})", decode_hex, text)
    text = re.sub(r"\\[a-zA-Z]+\d* ?", "", text)
    text = re.sub(r"\\[^ \n]+", "", text)
    text = text.replace("{", "").replace("}", "")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def _decode_rtf_line(raw_line: str) -> str:
    """Strip RTF tags khoi mot dong don, GIU NGUYEN dong trong."""
    def decode_hex(match):
        try:
            return bytes.fromhex(match.group(1)).decode("cp1252")
        except Exception:
            return ""

    line = re.sub(r"\\'([0-9a-fA-F]{2})", decode_hex, raw_line)
    line = re.sub(r"\\[a-zA-Z]+\d* ?", "", line)
    line = re.sub(r"\\[^ \n]+", "", line)
    line = line.replace("{", "").replace("}", "")
    line = re.sub(r"[ \t]{2,}", " ", line)
    return line.strip()


def _split_raw_batches(raw_rtf: str) -> list:
    """
    Tach raw RTF thanh cac batch tho (giu nguyen RTF tags).
    Tra ve list cac raw batch string, bat dau tu 'Batch Run 1:'.
    """
    parts = re.split(r"(?=Batch Run \d+:)", raw_rtf)
    return [p for p in parts if re.match(r"Batch Run \d+:", p.lstrip())]


def extract_number(line: str):
    match = re.search(NUMBER_PATTERN, line)
    return float(match.group()) if match else None


def extract_numbers(line: str):
    return [float(value) for value in re.findall(NUMBER_PATTERN, line)]


def parse_case_metadata(batch_lines):
    metadata = {
        "Water Level": None,
        "Draft": None,
        "Wind Speed": None,
        "Current": None,
        "Trim": None,
        "Bottom Clearance": None,
        "Current Direction": None,
        "Wind Direction": None,
    }

    for line in batch_lines:
        if "Water Level:" in line:
            metadata["Water Level"] = extract_number(line)
        elif "Draft:" in line:
            metadata["Draft"] = extract_number(line)
        elif "Wind Speed:" in line:
            metadata["Wind Speed"] = extract_number(line)
        elif re.match(r"^\s*Current:\s*", line):
            metadata["Current"] = extract_number(line)
        elif "Trim:" in line:
            metadata["Trim"] = extract_number(line)
        elif "Bottom Clearance:" in line:
            metadata["Bottom Clearance"] = extract_number(line)
        elif "Current Direction from:" in line:
            metadata["Current Direction"] = line.split("from:", 1)[1].strip()
        elif "Wind Direction from:" in line:
            metadata["Wind Direction"] = line.split("from:", 1)[1].strip()

    return metadata


def parse_line_tensions(batch_lines):
    rows = []
    in_block = False

    for raw_line in batch_lines:
        line = raw_line.strip()
        if "Line to" in line:
            in_block = True
            continue

        if not in_block:
            continue

        if "____" in line:
            break
        if not line:
            continue
        if line.startswith(("Bollard", "Pull", "Tot.Line", "In-Line", "Winch", "Worst", "Percent", "Strength")):
            continue

        parts = line.split()
        if len(parts) < 2:
            continue

        line_name = parts[0]
        numbers = extract_numbers(line)

        if numbers and re.match(r"^(\d+)[^\s]*", line_name):
            line_label_match = re.match(r"^(\d+)", line_name)
            if line_label_match and len(numbers) > 1 and int(numbers[0]) == int(line_label_match.group(1)):
                numbers = numbers[1:]

        if len(numbers) < 6:
            continue

        pull_in = numbers[0]
        tot_line_length = numbers[1]
        in_line_motion = None
        winch_slip = None
        worst_to_screen = None
        worst_to_true = None
        line_tension = None
        percent_strength = None

        if len(numbers) == 6:
            worst_to_screen = int(numbers[2])
            worst_to_true = int(numbers[3])
            line_tension = numbers[4]
            percent_strength = int(numbers[5])
        elif len(numbers) == 7:
            in_line_motion = numbers[2]
            worst_to_screen = int(numbers[3])
            worst_to_true = int(numbers[4])
            line_tension = numbers[5]
            percent_strength = int(numbers[6])

            if abs(in_line_motion) > 180:
                winch_slip = None
                in_line_motion = None
                worst_to_screen = int(numbers[2])
                worst_to_true = int(numbers[3])
                line_tension = numbers[4]
                percent_strength = int(numbers[5])
        else:
            in_line_motion = numbers[2]
            winch_slip = numbers[3]
            worst_to_screen = int(numbers[4])
            worst_to_true = int(numbers[5])
            line_tension = numbers[6]
            percent_strength = int(numbers[7])

        rows.append({
            "Line": line_name,
            "Pull-in": pull_in,
            "Tot. Line Length": tot_line_length,
            "In-Line Motion": in_line_motion,
            "Winch Slip": winch_slip,
            "Worst Direction to Screen": worst_to_screen,
            "Worst Direction to True North": worst_to_true,
            "Line Tension": line_tension,
            "Percent Strength": percent_strength,
        })

    return rows


def parse_fenders(batch_lines):
    rows = []
    in_block = False

    for raw_line in batch_lines:
        line = raw_line.strip()
        if "Fender" in line and "Thrust" in line:
            in_block = True
            continue

        if not in_block:
            continue

        if "____" in line:
            break
        if not line:
            continue
        if line.startswith(("Fender", "Hook/", "Bollard", "Total")):
            continue

        parts = line.split()
        if not parts:
            continue

        name = parts[0]
        nums = extract_numbers(line)
        if not nums:
            continue

        thrust = nums[0]
        compression = nums[1] if len(nums) > 1 else None
        pressure = nums[2] if len(nums) > 2 else None
        flatside_area = parts[-1] if parts[-1].endswith("%") else None

        rows.append({
            "Fender": name,
            "Thrust": thrust,
            "Compression": compression,
            "Pressure": pressure,
            "Flatside Area": flatside_area,
        })

    return rows


def _parse_greatest_section(batch_lines, title):
    rows = []
    in_block = False
    header_lines_to_skip = 0

    for raw_line in batch_lines:
        line = raw_line.strip()
        if title in line:
            in_block = True
            rows = []  # RESET: bo data lan truoc, chi giu lan cuoi
            header_lines_to_skip = 2
            continue

        if not in_block:
            continue

        if header_lines_to_skip > 0:
            header_lines_to_skip -= 1
            continue

        if not line:
            continue

        if (line.startswith("Greatest ") or line.startswith("Batch File:")
                or line.startswith("Batch Run") or line.startswith("Remarks:")):
            break

        tokens = [t for t in line.split() if t != "\x00"]
        if not tokens:
            continue

        all_numbers = extract_numbers(line)
        identifier = tokens[0]

        try:
            id_as_int = int(float(identifier))
            if all_numbers and int(all_numbers[0]) == id_as_int:
                numbers = all_numbers[1:]
            else:
                numbers = all_numbers
        except (ValueError, TypeError):
            numbers = all_numbers

        if len(numbers) < 9:
            continue

        primary_value     = numbers[0]
        wind_speed        = numbers[1]
        wind_direction    = numbers[2]
        current_speed     = numbers[3]
        current_direction = numbers[4]

        batch_no = int(numbers[-1])
        offset   = numbers[-2]
        trim     = numbers[-3]
        draft    = numbers[-4]

        middle = numbers[5:-4]

        wave_ht             = None
        wave_true_direction = None
        wave_period         = None
        water_level         = None

        if len(middle) == 4:
            wave_ht             = middle[0]
            wave_true_direction = middle[1]
            wave_period         = middle[2]
            water_level         = middle[3]
        elif len(middle) == 3:
            wave_ht     = middle[0]
            wave_period = middle[1]
            water_level = middle[2]
        elif len(middle) == 2:
            wave_ht     = middle[0]
            water_level = middle[1]
        elif len(middle) == 1:
            water_level = middle[0]

        rows.append({
            "_identifier":              identifier,
            "_primary_value":           primary_value,
            "Wind Speed":               wind_speed,
            "Wind Screen Direction":    wind_direction,
            "Current Speed":            current_speed,
            "Current Screen Direction": current_direction,
            "Wave Ht":                  wave_ht,
            "Wave True Direction":      wave_true_direction,
            "Wave Period":              wave_period,
            "Water Level":              water_level,
            "Draft":                    draft,
            "Trim":                     trim,
            "Offset":                   offset,
            "Batch Run no":             batch_no,
        })

    return rows


def parse_greatest_excursions(batch_lines):
    raw = _parse_greatest_section(batch_lines, "Greatest Excursions at Target:")
    rows = []
    for r in raw:
        rows.append({
            "Excursion Type":           r["_identifier"],
            "Highest Excursion":        r["_primary_value"],
            "Wind Speed":               r["Wind Speed"],
            "Wind Screen Direction":    r["Wind Screen Direction"],
            "Current Speed":            r["Current Speed"],
            "Current Screen Direction": r["Current Screen Direction"],
            "Wave Ht":                  r["Wave Ht"],
            "Wave True Direction":      r["Wave True Direction"],
            "Wave Period":              r["Wave Period"],
            "Water Level":              r["Water Level"],
            "Draft":                    r["Draft"],
            "Trim":                     r["Trim"],
            "Offset":                   r["Offset"],
            "Batch Run no":             r["Batch Run no"],
        })
    return rows


def parse_greatest_line_tensions(batch_lines):
    raw = _parse_greatest_section(batch_lines, "Greatest Line Tensions as % of Strength:")
    rows = []
    for r in raw:
        rows.append({
            "Line":                     r["_identifier"],
            "Highest Loading":          r["_primary_value"] / 100,
            "Wind Speed":               r["Wind Speed"],
            "Wind Screen Direction":    r["Wind Screen Direction"],
            "Current Speed":            r["Current Speed"],
            "Current Screen Direction": r["Current Screen Direction"],
            "Wave Ht":                  r["Wave Ht"],
            "Wave True Direction":      r["Wave True Direction"],
            "Wave Period":              r["Wave Period"],
            "Water Level":              r["Water Level"],
            "Draft":                    r["Draft"],
            "Trim":                     r["Trim"],
            "Offset":                   r["Offset"],
            "Batch Run no":             r["Batch Run no"],
        })
    return rows


def parse_greatest_berth_fender_thrusts(batch_lines):
    raw = _parse_greatest_section(batch_lines, "Greatest Berth Fender Thrusts:")
    rows = []
    for r in raw:
        rows.append({
            "Fender":                   r["_identifier"],
            "Thrust":                   r["_primary_value"],
            "Wind Speed":               r["Wind Speed"],
            "Wind Screen Direction":    r["Wind Screen Direction"],
            "Current Speed":            r["Current Speed"],
            "Current Screen Direction": r["Current Screen Direction"],
            "Wave Ht":                  r["Wave Ht"],
            "Wave True Direction":      r["Wave True Direction"],
            "Wave Period":              r["Wave Period"],
            "Water Level":              r["Water Level"],
            "Draft":                    r["Draft"],
            "Trim":                     r["Trim"],
            "Offset":                   r["Offset"],
            "Batch Run no":             r["Batch Run no"],
        })
    return rows


def parse_greatest_horizontal_bollard_forces(batch_lines):
    raw = _parse_greatest_section(batch_lines, "Greatest Horizontal Bollard Forces:")
    rows = []
    for r in raw:
        rows.append({
            "Bollard":                  r["_identifier"],
            "Force":                    r["_primary_value"],
            "Wind Speed":               r["Wind Speed"],
            "Wind Screen Direction":    r["Wind Screen Direction"],
            "Current Speed":            r["Current Speed"],
            "Current Screen Direction": r["Current Screen Direction"],
            "Wave Ht":                  r["Wave Ht"],
            "Wave True Direction":      r["Wave True Direction"],
            "Wave Period":              r["Wave Period"],
            "Water Level":              r["Water Level"],
            "Draft":                    r["Draft"],
            "Trim":                     r["Trim"],
            "Offset":                   r["Offset"],
            "Batch Run no":             r["Batch Run no"],
        })
    return rows


def parse_hook_bollard_forces(raw_batch_rtf: str) -> list:
    """
    Parse Hook/Bollard Forces tu raw RTF cua mot batch.

    Ly do dung raw RTF thay vi batch_lines (text da strip):
    - Cac bollard khong co luc (dong trong) bi xoa boi rtf_to_text()
      do buoc re.sub(r'\\n{2,}', '\\n', text).
    - Can giu nguyen dong trong de dem dung vi tri bollard A, B, C, ...
    - Parse theo tung \\par trong raw RTF dam bao moi dong RTF
      (ke ca trong) tuong ung dung mot bollard.

    Cau truc block trong RTF:
      Bollard A-Z  : khong co label, chi co so (hoac trong)
      Bollard *A-*Z: co label *A, *B, ... (hoac chi co label neu trong)
    """
    bollard_names = [chr(code) for code in range(ord("A"), ord("Z") + 1)]
    bollard_names += [f"*{chr(code)}" for code in range(ord("A"), ord("Z") + 1)]

    # Buoc 1: Tach raw RTF thanh cac dong theo \par, giu dong trong
    par_segments = re.split(r"\\par\b", raw_batch_rtf)
    decoded_lines = [_decode_rtf_line(seg) for seg in par_segments]

    # Buoc 2: Tim block Hook/Bollard
    block_start = None
    header_idx  = None
    block_end   = None

    for i, line in enumerate(decoded_lines):
        if block_start is None and "Hook/" in line and "Bollard" in line:
            block_start = i
            continue
        if block_start is not None and header_idx is None:
            if "Force" in line and "Strength" in line:
                header_idx = i
            continue
        if block_start is not None and header_idx is not None:
            if "____" in line:
                block_end = i
                break

    if block_start is None or header_idx is None:
        return []

    data_lines = decoded_lines[header_idx + 1 : block_end]

    # Buoc 3: Parse tung dong du lieu
    rows        = []
    current_idx = 0  # con tro trong bollard_names

    for line in data_lines:
        tokens = line.split()
        name   = None

        # Dong co label bollard (*A, *B, ... hoac A, B, ...)
        if tokens and re.match(r"^\*?[A-Z]$", tokens[0]):
            name   = tokens[0]
            values = extract_numbers(" ".join(tokens[1:]))
        else:
            values = extract_numbers(line)

        if name is not None:
            # Bollard co label: nhay toi dung vi tri trong danh sach
            while current_idx < len(bollard_names) and bollard_names[current_idx] != name:
                rows.append({
                    "Hook/Bollard": bollard_names[current_idx],
                    "X-Force": None, "Y-Force": None, "Total Force": None,
                    "% Strength": None, "Direction": None, "Uplift": None,
                })
                current_idx += 1
            current_name = name
        else:
            # Bollard khong co label: lay theo thu tu
            current_name = bollard_names[current_idx] if current_idx < len(bollard_names) else None

        current_idx += 1

        rows.append({
            "Hook/Bollard": current_name,
            "X-Force":      values[0]        if len(values) >= 1 else None,
            "Y-Force":      values[1]        if len(values) >= 2 else None,
            "Total Force":  values[2]        if len(values) >= 3 else None,
            "% Strength":   values[3] / 100  if len(values) >= 4 else None,
            "Direction":    int(values[4])   if len(values) >= 5 else None,
            "Uplift":       values[5]        if len(values) >= 6 else None,
        })

    # Dien cac bollard con lai chua xuat hien
    while current_idx < len(bollard_names):
        rows.append({
            "Hook/Bollard": bollard_names[current_idx],
            "X-Force": None, "Y-Force": None, "Total Force": None,
            "% Strength": None, "Direction": None, "Uplift": None,
        })
        current_idx += 1

    return rows


def reorder_columns(df, first_cols):
    cols = [c for c in first_cols if c in df.columns]
    cols += [c for c in df.columns if c not in cols]
    return df[cols]


def apply_number_formats(workbook, sheet_name, df, fmt_map):
    """Apply openpyxl number formats given {column_name: format_string}."""
    if sheet_name not in workbook.sheetnames:
        return
    sheet = workbook[sheet_name]
    header_to_letter = {
        cell.value: get_column_letter(idx)
        for idx, cell in enumerate(sheet[1], start=1)
    }
    for col_name, fmt in fmt_map.items():
        if col_name not in header_to_letter:
            continue
        letter = header_to_letter[col_name]
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"{letter}{row}"]
            if cell.value is not None:
                cell.number_format = fmt


# Sheet name constants (all <= 31 characters)
SH_CASES     = "Cases"
SH_FENDERS   = "Fenders"
SH_LINE      = "Line_Tensions"
SH_BOLLARD   = "Hook_Bollard_Forces"
SH_G_EXCUR   = "Greatest_Excursions"
SH_G_LINE    = "Greatest_Line"
SH_G_FENDER  = "Greatest_Fender"
SH_G_BOLLARD = "Greatest_Bollard"


def _process_file(file, cases, line_tensions, fenders, hook_bollard_forces,
                  greatest_excursions, greatest_line_tensions,
                  greatest_berth_fender_thrusts, greatest_horizontal_bollard_forces):
    """Doc va parse mot file RTF, append ket qua vao cac list truyen vao."""
    with open(file, "r", encoding="utf-8", errors="ignore") as f:
        raw_text = f.read()

    # Text-based batches (cho tat ca parser tru Hook/Bollard)
    text        = rtf_to_text(raw_text)
    batches     = text.split("Batch Run ")[1:]

    # Raw RTF batches (cho parse_hook_bollard_forces)
    raw_batches = _split_raw_batches(raw_text)

    for idx, batch in enumerate(batches):
        batch_id    = batch.split(":", 1)[0].strip()
        batch_lines = batch.splitlines()

        metadata = parse_case_metadata(batch_lines)
        cases.append({
            "File":              file.name,
            "Batch":             batch_id,
            "Water Level":       metadata["Water Level"],
            "Draft":             metadata["Draft"],
            "Wind Speed":        metadata["Wind Speed"],
            "Current":           metadata["Current"],
            "Trim":              metadata["Trim"],
            "Bottom Clearance":  metadata["Bottom Clearance"],
            "Current Direction": metadata["Current Direction"],
            "Wind Direction":    metadata["Wind Direction"],
        })

        line_tensions.extend([
            {**row, "File": file.name, "Batch": batch_id}
            for row in parse_line_tensions(batch_lines)
        ])
        fenders.extend([
            {**row, "File": file.name, "Batch": batch_id}
            for row in parse_fenders(batch_lines)
        ])

        # Hook/Bollard: dung raw RTF batch tuong ung
        raw_batch = raw_batches[idx] if idx < len(raw_batches) else ""
        hook_bollard_forces.extend([
            {**row, "File": file.name, "Batch": batch_id}
            for row in parse_hook_bollard_forces(raw_batch)
        ])

        greatest_excursions.extend(parse_greatest_excursions(batch_lines))
        greatest_line_tensions.extend(parse_greatest_line_tensions(batch_lines))
        greatest_berth_fender_thrusts.extend(parse_greatest_berth_fender_thrusts(batch_lines))
        greatest_horizontal_bollard_forces.extend(parse_greatest_horizontal_bollard_forces(batch_lines))


def _build_and_save(
    save_path,
    cases, line_tensions, fenders, hook_bollard_forces,
    greatest_excursions, greatest_line_tensions,
    greatest_berth_fender_thrusts, greatest_horizontal_bollard_forces,
    print_label="",
):
    """Xay dung DataFrames, ghi Excel, ap dung number formats."""
    df_cases         = pd.DataFrame(cases)
    df_fenders       = pd.DataFrame(fenders)
    df_line_tensions = pd.DataFrame(line_tensions)
    df_hook_bollard  = pd.DataFrame(hook_bollard_forces)
    df_greatest_excursions                = pd.DataFrame(greatest_excursions)
    df_greatest_line_tensions             = pd.DataFrame(greatest_line_tensions)
    df_greatest_berth_fender_thrusts      = pd.DataFrame(greatest_berth_fender_thrusts)
    df_greatest_horizontal_bollard_forces = pd.DataFrame(greatest_horizontal_bollard_forces)

    if "Percent Strength" in df_line_tensions.columns:
        df_line_tensions["Percent Strength"] = pd.to_numeric(
            df_line_tensions["Percent Strength"], errors="coerce"
        )
        df_line_tensions["Percent Strength"] = df_line_tensions["Percent Strength"].apply(
            lambda v: v / 100 if pd.notna(v) and v > 1 else v
        )

    first_cols = ["File", "Batch"]
    df_fenders       = reorder_columns(df_fenders,       first_cols)
    df_line_tensions = reorder_columns(df_line_tensions, first_cols)
    df_hook_bollard  = reorder_columns(df_hook_bollard,  first_cols)

    excursion_col_order = [
        "Excursion Type", "Highest Excursion",
        "Wind Speed", "Wind Screen Direction",
        "Current Speed", "Current Screen Direction",
        "Wave Ht", "Wave True Direction", "Wave Period",
        "Water Level", "Draft", "Trim", "Offset", "Batch Run no",
    ]
    if not df_greatest_excursions.empty:
        df_greatest_excursions = df_greatest_excursions[
            [c for c in excursion_col_order if c in df_greatest_excursions.columns]
        ]

    line_tension_col_order = [
        "Line", "Highest Loading",
        "Wind Speed", "Wind Screen Direction",
        "Current Speed", "Current Screen Direction",
        "Wave Ht", "Wave True Direction", "Wave Period",
        "Water Level", "Draft", "Trim", "Offset", "Batch Run no",
    ]
    if not df_greatest_line_tensions.empty:
        df_greatest_line_tensions = df_greatest_line_tensions[
            [c for c in line_tension_col_order if c in df_greatest_line_tensions.columns]
        ]

    fender_thrust_col_order = [
        "Fender", "Thrust",
        "Wind Speed", "Wind Screen Direction",
        "Current Speed", "Current Screen Direction",
        "Wave Ht", "Wave True Direction", "Wave Period",
        "Water Level", "Draft", "Trim", "Offset", "Batch Run no",
    ]
    if not df_greatest_berth_fender_thrusts.empty:
        df_greatest_berth_fender_thrusts = df_greatest_berth_fender_thrusts[
            [c for c in fender_thrust_col_order if c in df_greatest_berth_fender_thrusts.columns]
        ]

    bollard_force_col_order = [
        "Bollard", "Force",
        "Wind Speed", "Wind Screen Direction",
        "Current Speed", "Current Screen Direction",
        "Wave Ht", "Wave True Direction", "Wave Period",
        "Water Level", "Draft", "Trim", "Offset", "Batch Run no",
    ]
    if not df_greatest_horizontal_bollard_forces.empty:
        df_greatest_horizontal_bollard_forces = df_greatest_horizontal_bollard_forces[
            [c for c in bollard_force_col_order if c in df_greatest_horizontal_bollard_forces.columns]
        ]

    if print_label:
        print(f"\n{print_label}")
    print(f"  Cases:                            {len(df_cases)}")
    print(f"  Fenders:                          {len(df_fenders)}")
    print(f"  Line Tensions:                    {len(df_line_tensions)}")
    print(f"  Hook/Bollard Forces:              {len(df_hook_bollard)}")
    print(f"  Greatest Excursions:              {len(df_greatest_excursions)}")
    print(f"  Greatest Line Tensions:           {len(df_greatest_line_tensions)}")
    print(f"  Greatest Berth Fender Thrusts:    {len(df_greatest_berth_fender_thrusts)}")
    print(f"  Greatest Horizontal Bollard Forces:{len(df_greatest_horizontal_bollard_forces)}")

    def write_sheets(writer):
        df_cases.to_excel(writer,          sheet_name=SH_CASES,     index=False)
        df_fenders.to_excel(writer,        sheet_name=SH_FENDERS,   index=False)
        df_line_tensions.to_excel(writer,  sheet_name=SH_LINE,      index=False)
        df_hook_bollard.to_excel(writer,   sheet_name=SH_BOLLARD,   index=False)
        df_greatest_excursions.to_excel(writer,                sheet_name=SH_G_EXCUR,   index=False)
        df_greatest_line_tensions.to_excel(writer,             sheet_name=SH_G_LINE,    index=False)
        df_greatest_berth_fender_thrusts.to_excel(writer,     sheet_name=SH_G_FENDER,  index=False)
        df_greatest_horizontal_bollard_forces.to_excel(writer, sheet_name=SH_G_BOLLARD, index=False)

    actual_path = save_path
    try:
        with pd.ExcelWriter(actual_path, engine="openpyxl") as writer:
            write_sheets(writer)
    except PermissionError:
        stem = save_path.stem
        ts   = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        actual_path = save_path.with_name(f"{stem}_{ts}.xlsx")
        with pd.ExcelWriter(actual_path, engine="openpyxl") as writer:
            write_sheets(writer)

    # Number formats
    workbook = openpyxl.load_workbook(actual_path)

    _dir_fmt = '0'
    _pct_fmt = "0%"
    _num_fmt = "0.0"
    _int_fmt = "0"
    _exc_fmt = "0.00"

    apply_number_formats(workbook, SH_LINE, df_line_tensions, {
        "Percent Strength": _pct_fmt,
    })
    apply_number_formats(workbook, SH_BOLLARD, df_hook_bollard, {
        "% Strength":  _pct_fmt,
        "X-Force":     _num_fmt,
        "Y-Force":     _num_fmt,
        "Total Force": _num_fmt,
        "Uplift":      _num_fmt,
        "Direction":   _dir_fmt,
    })

    _gcf = {
        "Wind Speed":               _int_fmt,
        "Wind Screen Direction":    _dir_fmt,
        "Current Speed":            _num_fmt,
        "Current Screen Direction": _dir_fmt,
        "Wave Ht":                  _num_fmt,
        "Wave True Direction":      _dir_fmt,
        "Wave Period":              _num_fmt,
        "Water Level":              _num_fmt,
        "Draft":                    _num_fmt,
        "Trim":                     _num_fmt,
        "Offset":                   _num_fmt,
        "Batch Run no":             _int_fmt,
    }
    apply_number_formats(workbook, SH_G_EXCUR,   df_greatest_excursions,                {**_gcf, "Highest Excursion": _exc_fmt})
    apply_number_formats(workbook, SH_G_LINE,    df_greatest_line_tensions,             {**_gcf, "Highest Loading":   _pct_fmt})
    apply_number_formats(workbook, SH_G_FENDER,  df_greatest_berth_fender_thrusts,      {**_gcf, "Thrust":           _num_fmt})
    apply_number_formats(workbook, SH_G_BOLLARD, df_greatest_horizontal_bollard_forces, {**_gcf, "Force":            _num_fmt})

    workbook.save(actual_path)
    print(f"  Saved -> {actual_path.name}")


def main():
    # ── Per-file output ────────────────────────────────────────────────────────
    for file in rtf_files:
        cases                              = []
        line_tensions                      = []
        fenders                            = []
        hook_bollard_forces                = []
        greatest_excursions                = []
        greatest_line_tensions             = []
        greatest_berth_fender_thrusts      = []
        greatest_horizontal_bollard_forces = []

        _process_file(
            file, cases, line_tensions, fenders, hook_bollard_forces,
            greatest_excursions, greatest_line_tensions,
            greatest_berth_fender_thrusts, greatest_horizontal_bollard_forces,
        )

        _build_and_save(
            save_path=file.with_suffix(".xlsx"),
            cases=cases, line_tensions=line_tensions, fenders=fenders,
            hook_bollard_forces=hook_bollard_forces,
            greatest_excursions=greatest_excursions,
            greatest_line_tensions=greatest_line_tensions,
            greatest_berth_fender_thrusts=greatest_berth_fender_thrusts,
            greatest_horizontal_bollard_forces=greatest_horizontal_bollard_forces,
            print_label=f"[{file.name}]",
        )

    # ── Combined output ────────────────────────────────────────────────────────
    cases                              = []
    line_tensions                      = []
    fenders                            = []
    hook_bollard_forces                = []
    greatest_excursions                = []
    greatest_line_tensions             = []
    greatest_berth_fender_thrusts      = []
    greatest_horizontal_bollard_forces = []

    for file in rtf_files:
        _process_file(
            file, cases, line_tensions, fenders, hook_bollard_forces,
            greatest_excursions, greatest_line_tensions,
            greatest_berth_fender_thrusts, greatest_horizontal_bollard_forces,
        )

    _build_and_save(
        save_path=current_dir / "Optimoor_batch_results.xlsx",
        cases=cases, line_tensions=line_tensions, fenders=fenders,
        hook_bollard_forces=hook_bollard_forces,
        greatest_excursions=greatest_excursions,
        greatest_line_tensions=greatest_line_tensions,
        greatest_berth_fender_thrusts=greatest_berth_fender_thrusts,
        greatest_horizontal_bollard_forces=greatest_horizontal_bollard_forces,
        print_label="[Combined: Optimoor_batch_results.xlsx]",
    )


if __name__ == "__main__":
    main()
